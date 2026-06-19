"""
compare_results.py  —  comparative analysis of the pipeline's GeoDict results
against (a) Parente's reported kr/Sw data and (b) the paper's GeoDict absolute-
permeability reference.

Reads results.duckdb, computes relative permeabilities per (track, scan):
    kr_gas   = K_gas   / K_abs
    kr_brine = K_water / K_abs
joins Sw from fixed_boxes, and writes:
  - an Excel workbook with the per-domain table, the kr table, and a K_abs
    validation table
  - two figures: kr-vs-Sw (mine vs Parente) and K_abs-vs-reference

Parente's digitized curve is read from a CSV the user provides (see --parente).
The GeoDict absolute-permeability reference is 3.703e-13 m^2 (PC07/PC19 run).

Usage:
    python compare_results.py ^
        --db results.duckdb ^
        [--run-id <id>] [--connectivity 18N] ^
        [--parente parente_kr.csv] ^
        [--out comparison.xlsx]
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import duckdb

# Paper's GeoDict absolute-permeability reference (PC07/PC19 SolverResult_z.txt)
GEODICT_KABS_REF_M2 = 3.703e-13
M2_TO_DARCY = 1.0 / 9.869233e-13          # 1 Darcy = 9.869233e-13 m^2
M2_TO_MDARCY = M2_TO_DARCY * 1000.0


def fetch_kr_table(con, run_id, connectivity):
    """Per (track, scan): K_abs, K_gas, K_water, kr_gas, kr_brine, Sw.

    K_abs is the absolute (dry-rock) permeability for the track. It is solved at
    timestep X, so it may not be present at every scan; we take the per-track
    absolute value and apply it to every scan of that track.
    """
    conn_filter = "AND s.connectivity = ?" if connectivity else ""
    params = [run_id]
    if connectivity:
        params.append(connectivity)

    # gas and water permeabilities per (track, scan)
    rows = con.execute(f"""
        SELECT s.track_id, s.scan_index, s.connectivity, s.sim_type, s.k_z, s.Sw
        FROM simulation_results s
        WHERE s.run_id = ? {conn_filter}
          AND s.simulator = 'GeoDict'
        ORDER BY s.track_id, s.scan_index
    """, params).fetchall()

    # per-track Z position (box midpoint) for top-to-bottom ordering
    zrows = con.execute(f"""
        SELECT track_id, connectivity, MIN(z0), MIN(z1)
        FROM fixed_boxes
        WHERE run_id = ? {conn_filter.replace('s.', '')}
        GROUP BY track_id, connectivity
    """, params).fetchall()
    track_zmid = {}
    for track, conn, z0, z1 in zrows:
        if z0 is not None and z1 is not None:
            track_zmid[(track, conn)] = (z0 + z1) / 2.0
        elif z0 is not None:
            track_zmid[(track, conn)] = float(z0)

    # organise: per (track, conn) collect absolute; per (track, scan, conn) gas/water
    k_abs = {}                      # (track, conn) -> K_abs
    cell = {}                       # (track, scan, conn) -> {gas, water, sw}
    for track, scan, conn, sim_type, k_z, sw in rows:
        if sim_type == "absolute":
            # keep the absolute for the track (any scan it was solved at)
            k_abs[(track, conn)] = k_z
        else:
            d = cell.setdefault((track, scan, conn), {"gas": None, "water": None, "sw": None})
            if sim_type == "gas":
                d["gas"] = k_z
            elif sim_type == "water":
                d["water"] = k_z
            if sw is not None:
                d["sw"] = sw

    # Z rank: smaller box midpoint = nearer top/inlet = rank 1
    z_sorted = sorted(track_zmid, key=lambda k: track_zmid[k])
    z_rank = {k: i + 1 for i, k in enumerate(z_sorted)}

    out = []
    for (track, scan, conn), d in cell.items():
        kabs = k_abs.get((track, conn))
        kg, kw, sw = d["gas"], d["water"], d["sw"]
        kr_gas = (kg / kabs) if (kabs and kg is not None) else None
        kr_brine = (kw / kabs) if (kabs and kw is not None) else None
        out.append({
            "track_id": track, "scan_index": scan, "connectivity": conn,
            "z_rank": z_rank.get((track, conn), 999),
            "z_mid": track_zmid.get((track, conn)),
            "Sw": sw, "K_abs_m2": kabs, "K_gas_m2": kg, "K_water_m2": kw,
            "kr_gas": kr_gas, "kr_brine": kr_brine,
        })
    # order top-to-bottom by Z, then by scan
    out.sort(key=lambda r: (r["z_rank"], r["scan_index"]))
    return out, k_abs


def load_parente_xlsx(path: Path):
    """Read Parente's processed workbook ('Complete tables' sheet) and return
    her per-(section, scan) results: Sw, kr_gas (Rel perm H2), kr_brine
    (Rel perm brine), plus the per-section absolute permeability in mD.

    Comparison is on SATURATION, not on labels: Parente's 'Section' is a spatial
    sub-region and her 'Scan' numbering is sparse, neither of which maps onto the
    pipeline's tracks/scan indices. kr is a function of Sw, so the points are
    compared on the Sw axis regardless of how each side labelled regions/times.

    Returns (points, kabs_mD_list).
    """
    if not path or not path.exists():
        return None, None
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed; cannot read Parente workbook.")
        return None, None

    wb = openpyxl.load_workbook(path, data_only=True)
    if "Complete tables" not in wb.sheetnames:
        print("Parente workbook has no 'Complete tables' sheet.")
        return None, None
    ws = wb["Complete tables"]

    # Column layout (1-indexed) in 'Complete tables':
    #   5 Abs perm (GD) mD   7 Eff perm H2 (GD)   9 Eff perm brine (GD)
    #   11 Rel perm H2       12 Rel perm brine    13 Sg
    def _num(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    pts = []
    kabs_mD = []
    for r in range(5, ws.max_row + 1):
        scan = ws.cell(r, 4).value
        kabs = _num(ws.cell(r, 5).value)
        relh2 = _num(ws.cell(r, 11).value)
        relbr = _num(ws.cell(r, 12).value)
        sg = _num(ws.cell(r, 13).value)
        if kabs is not None:
            kabs_mD.append(kabs)
        # fluid scans (scan != 0) carry the kr points
        if scan not in (0, None) and sg is not None:
            pts.append({
                "Sw": 1.0 - sg,
                "kr_gas": relh2,
                "kr_brine": relbr,
            })
    return pts, kabs_mD


def match_nearest_sw(kr_rows, parente):
    """For each of MY (track, scan) points, find Parente's single nearest-Sw
    point and report the kr deviations. 1-to-1 nearest matching on Sw."""
    if not parente:
        return []
    ppts = [p for p in parente if p["Sw"] is not None]
    matched = []
    for r in kr_rows:
        if r["Sw"] is None:
            continue
        # nearest Parente point by |dSw|
        nearest = min(ppts, key=lambda p: abs(p["Sw"] - r["Sw"]))
        dsw = r["Sw"] - nearest["Sw"]
        def dev(mine, hers):
            if mine is None or hers is None or hers == 0:
                return None
            return (mine - hers) / hers * 100.0
        matched.append({
            "z_rank": r["z_rank"], "track_id": r["track_id"],
            "scan_index": r["scan_index"], "my_Sw": r["Sw"],
            "parente_Sw": nearest["Sw"], "dSw": dsw,
            "my_kr_gas": r["kr_gas"], "parente_kr_gas": nearest["kr_gas"],
            "kr_gas_dev_pct": dev(r["kr_gas"], nearest["kr_gas"]),
            "my_kr_brine": r["kr_brine"], "parente_kr_brine": nearest["kr_brine"],
            "kr_brine_dev_pct": dev(r["kr_brine"], nearest["kr_brine"]),
        })
    return matched
    ap.add_argument("--db",
                    default=r"C:\Users\99619\Desktop\SVETA\Micro-CT-Analysis-Pipeline\results.duckdb")
    ap.add_argument("--run-id", default=None, help="default: latest run in the DB")
    ap.add_argument("--connectivity", default=None, help="e.g. 18N; default: all")
    ap.add_argument("--parente",
                    default=r"C:\Users\99619\Desktop\SVETA\Micro-CT-Analysis-Pipeline\Methane_-_Data_processing.xlsx",
                    help="Parente's processed workbook (.xlsx) for the kr/Sw overlay")
    ap.add_argument("--out", default="comparison.xlsx")
    args = ap.parse_args()

    con = duckdb.connect(args.db, read_only=True)

    run_id = args.run_id
    if run_id is None:
        r = con.execute("SELECT run_id FROM runs ORDER BY rowid DESC LIMIT 1").fetchone()
        if not r:
            sys.exit("no runs in the DB.")
        run_id = r[0]
        print(f"using latest run_id: {run_id}")

    kr_rows, k_abs = fetch_kr_table(con, run_id, args.connectivity)
    con.close()
    if not kr_rows:
        sys.exit("no GeoDict results found for this run. Has the batch written any rows?")

    parente, parente_kabs_mD = (load_parente_xlsx(Path(args.parente))
                                if args.parente else (None, None))

    # ---- console summary (Z-ordered, top to bottom) ----
    print(f"\n{len(kr_rows)} (track, scan) gas/water points (ordered top->bottom by Z)\n")
    print(f"{'Zrank':>5} {'track':>5} {'scan':>4} {'conn':>5} {'Sw':>6} "
          f"{'kr_gas':>9} {'kr_brine':>9}")
    for r in kr_rows:
        sw = f"{r['Sw']:.3f}" if r["Sw"] is not None else "  -  "
        kg = f"{r['kr_gas']:.4f}" if r["kr_gas"] is not None else "   -   "
        kw = f"{r['kr_brine']:.4f}" if r["kr_brine"] is not None else "   -   "
        print(f"{r['z_rank']:>5} {r['track_id']:>5} {r['scan_index']:>4} {r['connectivity']:>5} "
              f"{sw:>6} {kg:>9} {kw:>9}")

    # ---- per-timestep nearest-Sw matching to Parente ----
    matched = match_nearest_sw(kr_rows, parente)
    if matched:
        print(f"\nPer-point nearest-Sw match to Parente (deviation in kr):")
        print(f"{'Zrank':>5} {'trk':>3} {'scan':>4} {'mySw':>6} {'parSw':>6} "
              f"{'krg dev%':>9} {'krw dev%':>9}")
        for m in matched:
            kgd = f"{m['kr_gas_dev_pct']:+.1f}" if m["kr_gas_dev_pct"] is not None else "  -  "
            kwd = f"{m['kr_brine_dev_pct']:+.1f}" if m["kr_brine_dev_pct"] is not None else "  -  "
            print(f"{m['z_rank']:>5} {m['track_id']:>3} {m['scan_index']:>4} "
                  f"{m['my_Sw']:>6.3f} {m['parente_Sw']:>6.3f} {kgd:>9} {kwd:>9}")

    # ---- K_abs validation ----
    # Two references: the single PC07 GeoDict run (3.703e-13 m^2 = 375 mD) and
    # Parente's own per-section absolute permeabilities (her GeoDict values).
    par_lo = par_hi = par_mean = None
    if parente_kabs_mD:
        par_lo, par_hi = min(parente_kabs_mD), max(parente_kabs_mD)
        par_mean = sum(parente_kabs_mD) / len(parente_kabs_mD)
        print(f"\nParente's absolute permeability (GeoDict): "
              f"{par_lo:.0f}-{par_hi:.0f} mD, mean {par_mean:.0f} mD")
    print(f"\nK_abs vs GeoDict reference ({GEODICT_KABS_REF_M2:.3e} m^2 = "
          f"{GEODICT_KABS_REF_M2*M2_TO_MDARCY:.1f} mD):")
    kabs_rows = []
    for (track, conn), kabs in sorted(k_abs.items()):
        if kabs is None:
            continue
        kabs_mD = kabs * M2_TO_MDARCY
        dev = (kabs - GEODICT_KABS_REF_M2) / GEODICT_KABS_REF_M2 * 100.0
        dev_par = ((kabs_mD - par_mean) / par_mean * 100.0) if par_mean else None
        row = {"track_id": track, "connectivity": conn,
               "K_abs_m2": kabs, "K_abs_mD": kabs_mD,
               "ref_m2": GEODICT_KABS_REF_M2,
               "ref_mD": GEODICT_KABS_REF_M2 * M2_TO_MDARCY,
               "deviation_pct": dev,
               "parente_mean_mD": par_mean,
               "deviation_vs_parente_pct": dev_par}
        kabs_rows.append(row)
        extra = f", {dev_par:+.1f}% vs Parente mean" if dev_par is not None else ""
        print(f"  track {track} {conn}: {kabs_mD:.1f} mD, "
              f"{dev:+.1f}% vs PC07 ref{extra}")

    _write_excel(args.out, kr_rows, kabs_rows, parente, matched)
    _write_figures(kr_rows, kabs_rows, parente)
    print(f"\nwrote {args.out}, kr_vs_sw.png, kabs_validation.png")


def _write_excel(out_path, kr_rows, kabs_rows, parente, matched=None):
    try:
        from openpyxl import Workbook
    except ImportError:
        print("openpyxl not installed; skipping Excel. pip install openpyxl")
        return
    wb = Workbook()

    ws = wb.active
    ws.title = "kr_table"
    ws.append(["z_rank", "track_id", "scan_index", "connectivity", "Sw",
               "K_abs_m2", "K_gas_m2", "K_water_m2", "kr_gas", "kr_brine"])
    for r in kr_rows:
        ws.append([r["z_rank"], r["track_id"], r["scan_index"], r["connectivity"],
                   r["Sw"], r["K_abs_m2"], r["K_gas_m2"], r["K_water_m2"],
                   r["kr_gas"], r["kr_brine"]])

    if matched:
        wsm = wb.create_sheet("matched_to_Parente")
        wsm.append(["z_rank", "track_id", "scan_index", "my_Sw", "parente_Sw", "dSw",
                    "my_kr_gas", "parente_kr_gas", "kr_gas_dev_pct",
                    "my_kr_brine", "parente_kr_brine", "kr_brine_dev_pct"])
        for m in matched:
            wsm.append([m["z_rank"], m["track_id"], m["scan_index"], m["my_Sw"],
                        m["parente_Sw"], m["dSw"], m["my_kr_gas"], m["parente_kr_gas"],
                        m["kr_gas_dev_pct"], m["my_kr_brine"], m["parente_kr_brine"],
                        m["kr_brine_dev_pct"]])

    ws2 = wb.create_sheet("Kabs_validation")
    ws2.append(["track_id", "connectivity", "K_abs_m2", "K_abs_mD",
                "ref_m2", "ref_mD", "deviation_pct",
                "parente_mean_mD", "deviation_vs_parente_pct"])
    for r in kabs_rows:
        ws2.append([r["track_id"], r["connectivity"], r["K_abs_m2"], r["K_abs_mD"],
                    r["ref_m2"], r["ref_mD"], r["deviation_pct"],
                    r.get("parente_mean_mD"), r.get("deviation_vs_parente_pct")])

    if parente:
        ws3 = wb.create_sheet("Parente_reference")
        ws3.append(["Sw", "kr_gas", "kr_brine"])
        for p in parente:
            ws3.append([p["Sw"], p["kr_gas"], p["kr_brine"]])

    wb.save(out_path)


def _write_figures(kr_rows, kabs_rows, parente):
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        print("matplotlib not installed; skipping figures. pip install matplotlib")
        return

    # ---- kr vs Sw ----
    fig, ax = plt.subplots(figsize=(7, 5))
    mine = [(r["Sw"], r["kr_gas"], r["kr_brine"]) for r in kr_rows if r["Sw"] is not None]
    mine.sort(key=lambda t: t[0])
    if mine:
        sw = [m[0] for m in mine]
        ax.plot([m[0] for m in mine if m[1] is not None],
                [m[1] for m in mine if m[1] is not None],
                "o-", color="tab:orange", label="kr_gas (pipeline)")
        ax.plot([m[0] for m in mine if m[2] is not None],
                [m[2] for m in mine if m[2] is not None],
                "s-", color="tab:blue", label="kr_brine (pipeline)")
    if parente:
        pg = [(p["Sw"], p["kr_gas"]) for p in parente if p["Sw"] is not None and p["kr_gas"] is not None]
        pw = [(p["Sw"], p["kr_brine"]) for p in parente if p["Sw"] is not None and p["kr_brine"] is not None]
        if pg:
            ax.plot([x for x, _ in pg], [y for _, y in pg], "o--", color="orange",
                    alpha=0.5, label="kr_gas (Parente)")
        if pw:
            ax.plot([x for x, _ in pw], [y for _, y in pw], "s--", color="navy",
                    alpha=0.5, label="kr_brine (Parente)")
    ax.set_xlabel("Water saturation $S_w$")
    ax.set_ylabel("Relative permeability $k_r$")
    ax.set_title("Relative permeability vs saturation")
    ax.legend()
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig("kr_vs_sw.png", dpi=150)
    plt.close(fig)

    # ---- K_abs validation ----
    if kabs_rows:
        fig, ax = plt.subplots(figsize=(7, 5))
        labels = [f"t{r['track_id']} {r['connectivity']}" for r in kabs_rows]
        vals = [r["K_abs_mD"] for r in kabs_rows]
        ax.bar(labels, vals, color="tab:green", alpha=0.7, label="pipeline K_abs")
        ax.axhline(GEODICT_KABS_REF_M2 * M2_TO_MDARCY, color="red", ls="--",
                   label=f"GeoDict ref ({GEODICT_KABS_REF_M2*M2_TO_MDARCY:.0f} mD)")
        ax.set_ylabel("Absolute permeability (mD)")
        ax.set_title("Absolute permeability vs GeoDict reference")
        ax.legend()
        plt.xticks(rotation=45, ha="right")
        fig.tight_layout()
        fig.savefig("kabs_validation.png", dpi=150)
        plt.close(fig)


if __name__ == "__main__":
    main()
